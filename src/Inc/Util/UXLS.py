# VladVons@gmail.com
# 2019.12.30

import os
import re

## - xls
from xlrd import open_workbook

## -- xlsx
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

import xlsxwriter

## -- ods
#https://www.it-swarm.net/ru/python/kak-preobrazovat-elektronnye-tablicy-opendocument-v-pandu-dataframe/1041032955/
#import odf.opendocument
#import xml.parsers.expat
from pyexcel_ods import get_data

xlsxwriter

class TFields():
  pass


class TXls():
  def __init__(self):
    self.Clear()

  def Clear(self):
    self.Data = {}
    self.Info = []

  @staticmethod
  def ToFloat(aValue):
    try:
      Result = round(float(aValue), 2)
    except:
      Result = 0
    return Result

  @staticmethod
  def ToDigit(aValue):
    if (aValue):
      aValue = re.sub('[^0-9]', '', str(aValue))
    return aValue

  def Report(self, aMsg):
    self.Info.append(aMsg)
    print(aMsg)

  def AddItem(self, aCode, aName, aPrice):
    Item = self.Data.get(aCode)
    if (Item):
      self.Report('Duplicate:%s, %s, %s' % (aCode, aName, Item.get('Name')))
    else:
      self.Data[aCode] = {'Name': aName, 'Price': aPrice}

  def LoadFields(self, aSheet, aCode, aName, aPrice):
      self.Fields = TFields()
      self.Fields.Sheet = aSheet
      self.Fields.Code = column_index_from_string(aCode) - 1
      self.Fields.Name = column_index_from_string(aName) - 1
      self.Fields.Price = column_index_from_string(aPrice) - 1

  def LoadFile_xls(self, aFile):
    wb = open_workbook(aFile)

    if (self.Fields.Sheet):
      ws = wb.sheet_by_name(self.Fields.Sheet)
    else:
      ws = wb.sheet_by_index(0)

    self.Clear()
    for i in range(0, ws.nrows):
      Code = self.ToDigit(ws.cell(i, self.Fields.Code).value)
      if (Code):
        self.AddItem(Code, ws.cell(i, self.Fields.Name).value, ws.cell(i, self.Fields.Price).value)
    self.Report('File:%s, Sheet:%s, Records:%s' % (aFile, ws.name, ws.nrows))

  def LoadFile_xlsx(self, aFile):
    wb = load_workbook(filename = aFile, read_only = True, data_only = True)

    if (self.Fields.Sheet):
      ws = wb[self.Fields.Sheet]
    else:
      ws = wb.active

    self.Clear()
    for row in ws.rows:
      Code = self.ToDigit(row[self.Fields.Code].value)
      if (Code):
        self.AddItem(Code, row[self.Fields.Name].value, row[self.Fields.Price].value)
    self.Report('File:%s, Sheet:%s, Records:%s' % (aFile, ws.title, ws.max_row))

  #self.Fields indexes doesnt work correctly with compined cells
  def LoadFile_ods(self, aFile):
    Data = get_data(aFile)
    if (self.Fields.Sheet):
      Sheet = self.Fields.Sheet
    else:
      Sheet = list(Data.keys())[0]

    self.Clear()
    Items = Data.get(Sheet)
    for Item in Items:
      MaxIdx = max(self.Fields.Code, self.Fields.Name, self.Fields.Price)
      if (MaxIdx < len(Item)):
        Code = self.ToDigit(Item[self.Fields.Code])
        if (Code):
          self.AddItem(Code, Item[self.Fields.Name], Item[self.Fields.Price])
    self.Report('File:%s, Sheet:%s, Records:%s' % (aFile, Sheet, len(Items)))

  def LoadFile(self, aFile):
    Ext = os.path.splitext(aFile)[1].lower()
    if (Ext == '.xls'):
      self.LoadFile_xls(aFile)
    elif (Ext == '.xlsx'):
      self.LoadFile_xlsx(aFile)
    elif (Ext == '.ods'):
      self.LoadFile_ods(aFile)
    else:
      print('unknown format %s' % Ext)

  def Compare(self, aTXls):
    Result = []
    # find missed items and items with different prices
    for Code in self.Data:
      Price1 = self.Data[Code].get('Price', 0)
      if (Price1 is None):
        continue
      Price1 = self.ToFloat(Price1)

      Name   = self.Data[Code].get('Name')
      if (aTXls.Data.get(Code)):
        Price2 = self.ToFloat(aTXls.Data[Code].get('Price', 0))
        if (Price1 != Price2):
          Result.append([Code, Name, Price1, Price2, round(Price2 - Price1, 3), round(100 - Price1 / Price2 * 100, 2)])
      else:
          Result.append([Code, Name, Price1, 0, 0, 0])

    # find new items
    for Code in aTXls.Data:
      Name = aTXls.Data[Code].get('Name')
      if (not self.Data.get(Code)):
          Price2 = self.ToFloat(aTXls.Data[Code].get('Price', 0))
          Result.append([Code, Name, 0, Price2, 0, 0])
    return Result

  #@staticmethod
  #def Export_NotWork_openpyxl(aData, aFile):
  #  wb = Workbook()
  #  ws = wb.active
  #  Data = ('Code', 'Name', 'Price1', 'Price2', 'Diff', 'PCent')
  #  ws.append(Data)
  #  for Item in aData:
  #    ws.append(Item)
  #  wb.save(aFile)

  @staticmethod
  def Export(aData, aFile):
    wb = xlsxwriter.Workbook(aFile)
    ws = wb.add_worksheet()

    aData.insert(0, ['Code', 'Name', 'Price1', 'Price2', 'Diff', 'PCent'])
    for IdxRow, Row in enumerate(aData):
      for IdxCol, Col in enumerate(Row):
        ws.write(IdxRow, IdxCol, Col)
    wb.close()


class TBrend():
  def __init__(self):
    self.Arr = {}
    self.Arr['laktalis'] = {'Sheet': 'по брендам', 'Code': 'B', 'Name': 'J', 'Price': 'AH'}
    self.Arr['lusdorf'] =  {'Sheet': '',           'Code': 'C', 'Name': 'D', 'Price': 'J'}

  def Compare(self, aBrend, aFile1, aFile2, aFileOut):
    Brend = self.Arr.get(aBrend)
    Xls = TXls(aSheet = Brend['Sheet'], aCode = Brend['Code'], aName = Brend['Name'], aPrice = Brend['Price'])
    Data = Xls.Compare(aFile1, aFile2)
    Xls.Export(Data, aFileOut)


#Brend = TBrend()
#Brend.Compare('laktalis', 'lactalis-p1.xlsx', 'lactalis-p2b.xlsx', 'p3.xlsx')
#Brend.Compare('lusdorf',  'ld-1.xlsx',         'ld-2.xlsx', 'p3.xlsx')
#Brend.Compare('lusdorf',  'ld-1.xls',         'ld-2.xls', 'p3.xlsx')
#Brend.Compare('lusdorf',  'ld-1.ods',         'ld-2.ods', 'p3.xlsx')
